from tkinter import *
import openpyxl
import urllib.request
from PIL import Image

# GUI
root = Tk()
root.title("Chatbot")

BG_GRAY = "#ABB2B9"
BG_COLOR = "#750000"
TEXT_COLOR = "#EAECEE"

FONT = "Helvetica 14"
FONT_BOLD = "Helvetica 13 bold"

# path = "GymExercisesDataset.xlsx"
# wb_obj = openpyxl.load_workbook(path)
# sheet = wb_obj.active


muscles = ["abdominals", "abductors", "adductors", "biceps", "calves", "chest", "forearms", "glutes", "hamstrings",
           "lats", "lower back", "middle back", "neck", "quadriceps", "quads", "shoulders", "traps", "triceps"]

exercises = []
# Send function


def send():
    send = "You -> " + e.get()
    txt.insert(END, "\n" + send)

    user = e.get().lower()

    muscle = ""

    if (user == "hello"):
        txt.insert(END, "\n" + "Bot -> Hi there, how can I help?")
        e.delete(0, END)
        return

    elif (user == "hi" or user == "hii" or user == "hiiii"):
        txt.insert(END, "\n" + "Bot -> Hi there, what can I do for you?")
        e.delete(0, END)
        return

    elif (user == "how are you"):
        txt.insert(END, "\n" + "Bot -> fine! and you")
        e.delete(0, END)
        return

    elif (user == "fine" or user == "i am good" or user == "i am doing good"):
        txt.insert(END, "\n" + "Bot -> Great! how can I help you.")
        e.delete(0, END)
        return

    elif (user == "thanks" or user == "thank you" or user == "now its my time"):
        txt.insert(END, "\n" + "Bot -> My pleasure !")
        e.delete(0, END)
        return

    elif (user == "what do you sell" or user == "what kinds of items are there" or user == "have you something"):
        txt.insert(END, "\n" + "Bot -> We have coffee and tea")
        e.delete(0, END)
        return

    elif (user == "tell me a joke" or user == "tell me something funny" or user == "crack a funny line"):
        txt.insert(
            END, "\n" + "Bot -> What did the buffalo say when his son left for college? Bison.! ")
        e.delete(0, END)
        return

    elif (user == "goodbye" or user == "see you later" or user == "see yaa" or user == "bye"):
        txt.insert(END, "\n" + "Bot -> Have a nice day!")
        exit(0)
        return

    user = user.split()

    i = 0
    while (i < len(user)):
        if user[i] == "ab" or user[i] == "abs":
            muscle = "abdominals"
        if user[i] == "abdominals":
            muscle = "abdominals"
        if user[i] == "abductor":
            muscle = "abductors"
        if user[i] == "abductors":
            muscle = "abductors"
        if user[i] == "adductor":
            muscle = "adductors"
        if user[i] == "adductors":
            muscle = "adductors"
        if user[i] == "calve":
            muscle = "calves"
        if user[i] == "calves" or user[i] == "calf":
            muscle = "calves"
        if user[i] == "chest":
            muscle = "chest"
        if user[i] == "lower":
            if user[i+1] == "back":
                muscle = "lower back"
                i = i+1
        if user[i] == "middle":
            if user[i+1] == "back":
                muscle = "middle back"
                i = i+1
        if user[i] == "forearm":
            muscle = "forearms"
        if user[i] == "forearms":
            muscle = "forearms"
        if user[i] == "glute":
            muscle = "glutes"
        if user[i] == "glutes":
            muscle = "glutes"
        if user[i] == "hamstring":
            muscle = "hamstrings"
        if user[i] == "hamstrings":
            muscle = "hamstrings"
        if user[i] == "lat":
            muscle = "lats"
        if user[i] == "lats":
            muscle = "lats"
        if user[i] == "quad" or user[i] == "quads" or user[i] == "quadricep":
            muscle = "quadriceps"
        if user[i] == "quadriceps":
            muscle = "quadriceps"
        if user[i] == "shoulder":
            muscle = "shoulders"
        if user[i] == "shoulders":
            muscle = "shoulders"
        if user[i] == "trap":
            muscle = "traps"
        if user[i] == "traps":
            muscle = "traps"
        if user[i] == "tricep":
            muscle = "triceps"
        if user[i] == "triceps":
            muscle = "triceps"
        if user[i] == "biceps":
            muscle = "biceps"
        if user[i] == "bicep":
            muscle = "biceps"
        if user[i] == "neck":
            muscle = "neck"

        i = i+1

    # txt.insert(END, "\n" + "Bot ->" + muscle)

    if muscle == "":
        txt.insert(
            END, "\n" + "Bot -> Sorry, I'm not trained enough to answer that :/")
        e.delete(0, END)
        return

    # txt.insert(END, "\n" + "Bot -> Present")
    path = muscle+".xlsx"
    # print(path)
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj.active

    # txt.insert(END, "\n" + "Bot -> " + str(sheet[1][2].value))

    out = False

    for row in sheet.iter_rows(min_row=2, min_col=1):
        if (str(row[1].value) not in exercises):
            txt.insert(END, "\n" + "Bot -> " + str(row[1].value) + "\nEquipment: " + str(
                row[4].value) + "\nRating: " + str(row[5].value))
            e.delete(0, END)
            exercises.append(str(row[1].value))
            print(exercises)
            print(str(row[3].value))

            try:
                urllib.request.urlretrieve(str(row[2].value), "exercise.jpg")
            except:
                txt.insert(END, "\n" + "Bot -> Image not in database :/")
            else:
                urllib.request.urlretrieve(str(row[2].value), "exercise.jpg")
                img = Image.open("exercise.jpg")
                img.show()
                txt.insert(
                    END, "\n" + "Bot -> Image depicting the execution of exercise opened")
            out = True
            break

    if (not out):
        txt.insert(
            END, "\n" + "Bot -> Sorry, no more exercises of this muscle group in database :/")

    # if(str(row[1].value) == None):
    #         txt.insert(END, "\n" + "Bot -> No more exercises in the database for this muscle group :/")

    # for i in muscles:
    # 	if i in user:
    # 		# txt.insert(END, "\n" + "Bot -> Present")
    # 		n=1
    # for row in sheet:
    # 	# if (row[5].value == i):
    # 	if (row[5].value.lower()=="forearms"):
    # 		txt.insert(END, "\n" + "Bot ->" + str(row[0].value))
    # 		break

    e.delete(0, END)


lable1 = Label(root, bg=BG_COLOR, fg=TEXT_COLOR, text="Welcome", font=FONT_BOLD, pady=10, width=20, height=1).grid(
    row=0)

txt = Text(root, bg=BG_COLOR, fg=TEXT_COLOR, font=FONT, width=80)
txt.grid(row=1, column=0, columnspan=2)

scrollbar = Scrollbar(txt)
scrollbar.place(relheight=1, relx=0.98)

e = Entry(root, bg=BG_COLOR, fg=TEXT_COLOR, font=FONT, width=75)
e.grid(row=2, column=0)

send = Button(root, text="Send", font=FONT_BOLD, bg=BG_GRAY,
              command=send).grid(row=2, column=1)

txt.insert(END, "Welcome to the gym exercises chatbot. You can ask me anything related to exercises. The way I work is ask me for an exercise of a particular muscle group and I will give you the best exercise available in the database. Make sure to type in the correct spelling of the muscle group. If you want to close the chat, just type in 'bye' or 'goodbye' or just close the window. Enjoy!!\n\n")

# root.bind('<Return>', send)

root.mainloop()
